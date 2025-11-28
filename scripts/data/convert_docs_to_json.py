#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 docs 目录下的文件转换为 JSON 格式
"""

import os
import json
import sys
from pathlib import Path

# 添加项目根目录到路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

# 检查可用的 Excel 处理库
HAS_PANDAS = False
HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    pass

if not HAS_PANDAS:
    try:
        import openpyxl
        HAS_OPENPYXL = True
    except ImportError:
        print("警告: 未安装 pandas 或 openpyxl，无法处理 Excel 文件")


def convert_excel_to_json(excel_path, output_path):
    """将 Excel 文件转换为 JSON"""
    if not HAS_PANDAS and not HAS_OPENPYXL:
        print(f"跳过 Excel 文件 {excel_path}（缺少依赖库）")
        return False
    
    try:
        if HAS_PANDAS:
            # 使用 pandas 读取 Excel
            excel_data = {}
            xls = pd.ExcelFile(excel_path)
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                # 将 DataFrame 转换为字典列表
                # 处理 NaN 值
                df = df.fillna('')
                # 转换为字典列表
                excel_data[sheet_name] = df.to_dict('records')
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(excel_data, f, ensure_ascii=False, indent=2)
            
            print(f"✓ Excel 文件已转换: {excel_path} -> {output_path}")
            return True
        else:
            # 使用 openpyxl 读取 Excel
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_path, data_only=True)
            excel_data = {}
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                headers = None
                
                for row in sheet.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(cell) if cell is not None else '' for cell in row]
                    else:
                        row_data = {}
                        for i, cell in enumerate(row):
                            header = headers[i] if i < len(headers) else f'Column{i+1}'
                            row_data[header] = str(cell) if cell is not None else ''
                        rows.append(row_data)
                
                excel_data[sheet_name] = rows
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(excel_data, f, ensure_ascii=False, indent=2)
            
            print(f"✓ Excel 文件已转换: {excel_path} -> {output_path}")
            return True
    except Exception as e:
        print(f"✗ Excel 文件转换失败 {excel_path}: {e}")
        return False


def convert_text_to_json(file_path, output_path, file_type='text'):
    """将文本文件（Markdown、Python 等）转换为 JSON"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 创建 JSON 结构
        json_data = {
            'filename': os.path.basename(file_path),
            'file_type': file_type,
            'content': content
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        print(f"✓ {file_type} 文件已转换: {file_path} -> {output_path}")
        return True
    except Exception as e:
        print(f"✗ {file_type} 文件转换失败 {file_path}: {e}")
        return False


def convert_json_to_json(file_path, output_path):
    """将 JSON 文件重新格式化（如果需要）"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # 添加元数据
        enhanced_data = {
            'filename': os.path.basename(file_path),
            'file_type': 'json',
            'data': json_data
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(enhanced_data, f, ensure_ascii=False, indent=2)
        
        print(f"✓ JSON 文件已转换: {file_path} -> {output_path}")
        return True
    except Exception as e:
        print(f"✗ JSON 文件转换失败 {file_path}: {e}")
        return False


def main():
    """主函数"""
    docs_dir = os.path.join(project_root, 'docs')
    output_dir = os.path.join(project_root, 'docs', 'json_output')
    
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    
    print("=" * 60)
    print("开始转换 docs 目录下的文件为 JSON")
    print("=" * 60)
    
    if not os.path.exists(docs_dir):
        print(f"✗ docs 目录不存在: {docs_dir}")
        return
    
    # 获取所有文件
    files = []
    for file_name in os.listdir(docs_dir):
        file_path = os.path.join(docs_dir, file_name)
        if os.path.isfile(file_path) and not file_name.startswith('.'):
            files.append((file_name, file_path))
    
    print(f"\n找到 {len(files)} 个文件\n")
    
    success_count = 0
    skip_count = 0
    
    for file_name, file_path in files:
        # 跳过输出目录
        if 'json_output' in file_path:
            continue
        
        # 生成输出文件名
        base_name = os.path.splitext(file_name)[0]
        output_path = os.path.join(output_dir, f"{base_name}.json")
        
        # 根据文件类型转换
        ext = os.path.splitext(file_name)[1].lower()
        
        if ext == '.xlsx' or ext == '.xls':
            if convert_excel_to_json(file_path, output_path):
                success_count += 1
            else:
                skip_count += 1
        elif ext == '.md':
            if convert_text_to_json(file_path, output_path, 'markdown'):
                success_count += 1
            else:
                skip_count += 1
        elif ext == '.py':
            if convert_text_to_json(file_path, output_path, 'python'):
                success_count += 1
            else:
                skip_count += 1
        elif ext == '.json':
            if convert_json_to_json(file_path, output_path):
                success_count += 1
            else:
                skip_count += 1
        else:
            # 其他文本文件
            if convert_text_to_json(file_path, output_path, 'text'):
                success_count += 1
            else:
                skip_count += 1
    
    print("\n" + "=" * 60)
    print(f"转换完成！")
    print(f"成功: {success_count} 个文件")
    print(f"跳过/失败: {skip_count} 个文件")
    print(f"输出目录: {output_dir}")
    print("=" * 60)


if __name__ == '__main__':
    main()

