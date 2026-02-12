#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
年运报告批量导出脚本（独立版，无需项目代码）

从 Excel 读取用户信息（USER_ID、姓名、生辰、性别），
调用服务端 /annual-report/stream 接口生成年运报告，结果写回 Excel E 列。

LLM 平台（Coze/百炼）由服务端配置决定，无需本地配置 API Key，
确保参数和提示词与生产环境完全一致。

依赖安装:
    pip install aiohttp pandas openpyxl

用法:
    python3 annual_report_exporter.py \\
        --input /path/to/users.xlsx \\
        --year 2025 \\
        --concurrency 3 \\
        --verbose
"""

import os
import re
import json
import asyncio
import argparse
import time
import logging
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass

import pandas as pd
import aiohttp
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# ==================== Excel 列映射 ====================

class Columns:
    """Excel 列索引（基于0索引）"""
    USER_ID = 0         # A列: USER_ID
    USER_NAME = 1       # B列: 用户名
    USER_BIRTH = 2      # C列: 生辰（含时间，如 "1987年1月7日 9:15"）
    GENDER = 3          # D列: 性别（男/女）
    ANNUAL_REPORT = 4   # E列: 年运报告（输出）


# ==================== 数据模型 ====================

@dataclass
class UserRecord:
    """用户记录"""
    row_index: int          # pandas 行索引（从0开始）
    user_id: str
    user_name: str
    birth_text: str         # 原始生辰文本
    gender_text: str        # 原始性别文本
    solar_date: str         # 解析后的日期 YYYY-MM-DD
    solar_time: str         # 解析后的时间 HH:MM
    gender: str             # 解析后的性别 male/female
    parse_error: Optional[str] = None


# ==================== 生辰与性别解析 ====================

def parse_birth_with_time(text: str) -> Tuple[str, str]:
    """
    解析包含时间的生辰文本
    
    支持格式：
    - "1987年1月7日 9:15"  /  "1987年1月7日 09:15"
    - "1987 年 1 月 7 日 9:15"（带空格）
    - "1987-01-07 9:15"  /  "1987/01/07 9:15"
    - "1987年1月7日"（无时间则默认 12:00）
    
    Returns:
        (solar_date, solar_time) 元组
    """
    if not text or str(text).strip() == "" or str(text) == "nan":
        raise ValueError("生辰为空")
    
    text = str(text).strip()
    
    # 1. 提取日期部分
    date_patterns = [
        r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日',
        r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})',
    ]
    
    solar_date = None
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
            solar_date = f"{year:04d}-{month:02d}-{day:02d}"
            break
    
    if not solar_date:
        raise ValueError(f"无法解析日期: {text}")
    
    # 2. 提取时间部分（在 "日" 之后或日期之后查找 HH:MM）
    # 先尝试在 "日" 之后查找
    ri_match = re.search(r'日\s*(\d{1,2})\s*[：:]\s*(\d{2})', text)
    if ri_match:
        hour = int(ri_match.group(1))
        minute = ri_match.group(2)
        if 0 <= hour <= 23:
            return solar_date, f"{hour:02d}:{minute}"
    
    # 再尝试通用时间模式（文本末尾的 HH:MM）
    time_match = re.search(r'(\d{1,2})\s*[：:]\s*(\d{2})\s*$', text)
    if time_match:
        hour = int(time_match.group(1))
        minute = time_match.group(2)
        if 0 <= hour <= 23:
            return solar_date, f"{hour:02d}:{minute}"
    
    # 未找到时间，使用默认值
    return solar_date, "12:00"


def parse_gender(text: str) -> str:
    """
    解析性别文本
    
    Returns:
        "male" 或 "female"
    """
    if not text or str(text).strip() == "" or str(text) == "nan":
        raise ValueError("性别为空")
    
    text = str(text).strip().lower()
    
    if text in ['男', 'male', 'm', '1']:
        return 'male'
    elif text in ['女', 'female', 'f', '0', '2']:
        return 'female'
    else:
        raise ValueError(f"无法识别的性别: {text}")


# ==================== SSE 流式 HTTP 客户端 ====================

@dataclass
class StreamResponse:
    """流式响应结果"""
    content: str = ""
    error: Optional[str] = None


class ApiClient:
    """API 客户端（自包含，不依赖项目代码）"""
    
    def __init__(self, base_url: str = "http://8.210.52.217:8001",
                 stream_timeout: int = 300):
        self.base_url = base_url.rstrip('/')
        self.api_prefix = "/api/v1"
        self.stream_timeout = stream_timeout
        self._session: Optional[aiohttp.ClientSession] = None
    
    async def _get_session(self) -> aiohttp.ClientSession:
        if self._session is None or self._session.closed:
            timeout = aiohttp.ClientTimeout(total=self.stream_timeout, connect=30)
            self._session = aiohttp.ClientSession(timeout=timeout)
        return self._session
    
    async def close(self):
        if self._session and not self._session.closed:
            await self._session.close()
    
    async def _post_json(self, endpoint: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """发送 POST JSON 请求"""
        session = await self._get_session()
        url = f"{self.base_url}{self.api_prefix}{endpoint}"
        
        try:
            async with session.post(url, json=data) as response:
                response.raise_for_status()
                return await response.json()
        except aiohttp.ClientResponseError as e:
            raise RuntimeError(f"API 请求失败: {url}, {e.status}, {e.message}")
        except aiohttp.ClientError as e:
            raise RuntimeError(f"API 请求失败: {url}, {e}")
    
    async def _post_stream(self, endpoint: str, data: Dict[str, Any]) -> StreamResponse:
        """发送 POST 请求并解析 SSE 流式响应"""
        session = await self._get_session()
        url = f"{self.base_url}{self.api_prefix}{endpoint}"
        result = StreamResponse()
        progress_chunks = []
        buffer = b''
        
        try:
            async with session.post(url, json=data) as response:
                response.raise_for_status()
                
                try:
                    async for chunk in response.content.iter_chunked(8192):
                        if not chunk:
                            continue
                        
                        buffer += chunk
                        
                        while b'\n' in buffer:
                            line_bytes, buffer = buffer.split(b'\n', 1)
                            try:
                                line = line_bytes.decode('utf-8').strip()
                            except UnicodeDecodeError:
                                continue
                            
                            if not line or not line.startswith('data:'):
                                continue
                            
                            json_str = line[5:].strip()
                            if not json_str:
                                continue
                            
                            try:
                                msg = json.loads(json_str)
                                msg_type = msg.get('type', '')
                                
                                if msg_type == 'progress':
                                    progress_chunks.append(msg.get('content', ''))
                                elif msg_type == 'complete':
                                    result.content = msg.get('content', '')
                                    if not result.content and progress_chunks:
                                        result.content = ''.join(progress_chunks)
                                    return result
                                elif msg_type == 'error':
                                    result.error = msg.get('content', '未知错误')
                                    return result
                            except json.JSONDecodeError:
                                continue
                    
                    # 处理剩余 buffer
                    if buffer:
                        try:
                            line = buffer.decode('utf-8').strip()
                            if line.startswith('data:'):
                                json_str = line[5:].strip()
                                if json_str:
                                    msg = json.loads(json_str)
                                    if msg.get('type') == 'complete':
                                        result.content = msg.get('content', '')
                                    elif msg.get('type') == 'progress':
                                        progress_chunks.append(msg.get('content', ''))
                        except (UnicodeDecodeError, json.JSONDecodeError):
                            pass
                    
                    if not result.content and progress_chunks:
                        result.content = ''.join(progress_chunks)
                
                except (aiohttp.ClientPayloadError, aiohttp.ServerDisconnectedError):
                    if progress_chunks:
                        result.content = ''.join(progress_chunks) + " [传输中断]"
                        return result
                    raise
        
        except (aiohttp.ClientPayloadError, aiohttp.ServerDisconnectedError) as e:
            if progress_chunks:
                result.content = ''.join(progress_chunks) + " [传输中断]"
            else:
                result.error = f"传输错误: {e}"
        except aiohttp.ClientError as e:
            if progress_chunks:
                result.content = ''.join(progress_chunks) + " [连接中断]"
            else:
                result.error = f"请求失败: {e}"
        except asyncio.TimeoutError:
            if progress_chunks:
                result.content = ''.join(progress_chunks) + " [超时]"
            else:
                result.error = f"请求超时: {url}"
        except Exception as e:
            if progress_chunks:
                result.content = ''.join(progress_chunks) + f" [异常: {type(e).__name__}]"
            else:
                result.error = f"请求异常: {type(e).__name__}: {e}"
        
        if not result.content and not result.error:
            result.error = f"API 返回空内容: {url}"
        
        return result
    
    async def _post_stream_with_retry(self, endpoint: str, data: Dict[str, Any],
                                       max_retries: int = 3) -> StreamResponse:
        """带重试的流式请求"""
        last_error = None
        best_result = None
        
        for attempt in range(max_retries):
            result = await self._post_stream(endpoint, data)
            
            # 成功获取足够内容
            if result.content and len(result.content) >= 100 and not result.error:
                return result
            
            # 保存最佳结果
            if result.content and len(result.content) >= 100:
                if best_result is None or len(result.content) > len(best_result.content):
                    best_result = result
            
            if result.error:
                last_error = result.error
                # 非传输错误不重试
                should_retry = any(kw in result.error for kw in
                                   ['传输错误', '传输中断', '连接中断', '超时', '空内容'])
                if not should_retry:
                    return result
            
            if attempt < max_retries - 1:
                wait_time = 2 ** attempt
                await asyncio.sleep(wait_time)
        
        if best_result and best_result.content:
            return best_result
        
        result = StreamResponse()
        result.error = f"接口调用失败（已重试 {max_retries} 次）: {last_error or '未知错误'}"
        return result
    
    async def call_annual_report_stream(self, solar_date: str, solar_time: str,
                                         gender: str, year: Optional[int] = None,
                                         app_id: Optional[str] = None) -> StreamResponse:
        """
        调用年运报告流式接口。
        与生产环境调用的是同一个端点；app_id 可选，用于指定百炼智能体（如 2025 年新提示词）。
        """
        data = {
            "solar_date": solar_date,
            "solar_time": solar_time,
            "gender": gender
        }
        if year is not None:
            data["year"] = year
        if app_id:
            data["app_id"] = app_id
        return await self._post_stream_with_retry("/annual-report/stream", data, max_retries=3)
    

# ==================== Excel 处理器 ====================

class ExcelHandler:
    """Excel 读写处理器"""
    
    def __init__(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Excel 文件不存在: {filepath}")
        
        self.filepath = filepath
        self._workbook = load_workbook(filepath)
        self._sheet = self._workbook.active
        self._df = pd.read_excel(filepath, header=0)
    
    def read_user_records(self, data_row_start: Optional[int] = None,
                          data_row_end: Optional[int] = None) -> List[UserRecord]:
        """
        读取用户记录
        
        Args:
            data_row_start: 起始数据行号（从1开始）
            data_row_end: 结束数据行号（包含）
        """
        records = []
        
        for idx, row in self._df.iterrows():
            data_row_num = idx + 1
            
            if data_row_start is not None and data_row_num < data_row_start:
                continue
            if data_row_end is not None and data_row_num > data_row_end:
                break
            
            user_id = str(row.iloc[Columns.USER_ID]) if pd.notna(row.iloc[Columns.USER_ID]) else str(idx + 1)
            user_name = str(row.iloc[Columns.USER_NAME]) if pd.notna(row.iloc[Columns.USER_NAME]) else ""
            birth_text = str(row.iloc[Columns.USER_BIRTH]) if pd.notna(row.iloc[Columns.USER_BIRTH]) else ""
            gender_text = str(row.iloc[Columns.GENDER]) if pd.notna(row.iloc[Columns.GENDER]) else ""
            
            if not birth_text or birth_text == "nan":
                continue
            
            record = UserRecord(
                row_index=idx,
                user_id=user_id,
                user_name=user_name,
                birth_text=birth_text,
                gender_text=gender_text,
                solar_date="",
                solar_time="",
                gender=""
            )
            
            try:
                record.solar_date, record.solar_time = parse_birth_with_time(birth_text)
            except ValueError as e:
                record.parse_error = f"生辰解析失败: {e}"
            
            try:
                record.gender = parse_gender(gender_text)
            except ValueError as e:
                if record.parse_error:
                    record.parse_error += f"; 性别解析失败: {e}"
                else:
                    record.parse_error = f"性别解析失败: {e}"
            
            records.append(record)
        
        return records
    
    def write_result(self, record: UserRecord, content: str):
        """写入年运报告结果到 E 列"""
        excel_row = record.row_index + 2  # 1行表头 + openpyxl从1开始
        self._sheet.cell(row=excel_row, column=Columns.ANNUAL_REPORT + 1, value=content)
    
    def save(self):
        self._workbook.save(self.filepath)
    
    def close(self):
        if self._workbook:
            self._workbook.close()


# ==================== 核心导出逻辑 ====================

class AnnualReportExporter:
    """年运报告批量导出器
    
    统一通过 /annual-report/stream 调用服务端。可选传入 app_id 指定百炼智能体
    （如 2025 年新提示词），不传则使用服务端数据库配置。
    """
    
    def __init__(self, base_url: str, year: int, app_id: Optional[str] = None, verbose: bool = False):
        self.year = year
        self.app_id = app_id
        self.verbose = verbose
        self.api_client = ApiClient(base_url=base_url)
    
    async def _process_stream(self, record: UserRecord) -> Optional[str]:
        """调用 /annual-report/stream"""
        try:
            result = await self.api_client.call_annual_report_stream(
                solar_date=record.solar_date,
                solar_time=record.solar_time,
                gender=record.gender,
                year=self.year,
                app_id=self.app_id
            )
            if result.error:
                logger.error(f"[{record.user_name}] 失败: {result.error}")
                return f"[错误] {result.error}"
            if result.content:
                return result.content
            return "[错误] 返回空内容"
        except Exception as e:
            logger.error(f"[{record.user_name}] 异常: {e}")
            return f"[异常] {e}"
    
    async def _process_user(self, record: UserRecord, semaphore: asyncio.Semaphore,
                             excel: ExcelHandler, progress: Dict[str, int]):
        """处理单个用户"""
        async with semaphore:
            label = f"[{record.user_id}] {record.user_name}"
            
            if record.parse_error:
                print(f"  跳过: {label} - {record.parse_error}")
                progress['failed'] += 1
                return
            
            if self.verbose:
                print(f"  处理中: {label} ({record.solar_date} {record.solar_time}, "
                      f"{'男' if record.gender == 'male' else '女'})")
            
            start = time.time()
            content = await self._process_stream(record)
            elapsed = time.time() - start
            
            # 写入 E 列
            excel.write_result(record, content=content)
            excel.save()
            
            ok = content and not content.startswith("[错误]") and not content.startswith("[异常]")
            if ok:
                progress['success'] += 1
                if self.verbose:
                    print(f"  完成: {label} ({elapsed:.1f}s) - {content[:80]}...")
            else:
                progress['failed'] += 1
                print(f"  失败: {label} ({elapsed:.1f}s) - {(content or '未知错误')[:100]}")
    
    async def export(self, input_path: str, max_concurrency: int = 2,
                     row: Optional[int] = None):
        """批量导出年运报告"""
        print("=" * 60)
        print("年运报告批量导出")
        print("=" * 60)
        print(f"  输入文件: {input_path}")
        print(f"  目标年份: {self.year}")
        if self.app_id:
            print(f"  百炼智能体: {self.app_id}")
        print(f"  并发数: {max_concurrency}")
        print(f"  API地址: {self.api_client.base_url}")
        print("=" * 60)
        
        excel = ExcelHandler(input_path)
        
        data_row_start = row if row else None
        data_row_end = row if row else None
        records = excel.read_user_records(data_row_start, data_row_end)
        
        if not records:
            print("\n未找到有效的用户记录")
            excel.close()
            return
        
        print(f"\n共 {len(records)} 条用户记录待处理\n")
        
        semaphore = asyncio.Semaphore(max_concurrency)
        progress = {'success': 0, 'failed': 0}
        start_time = time.time()
        
        tasks = [
            self._process_user(record, semaphore, excel, progress)
            for record in records
        ]
        await asyncio.gather(*tasks)
        
        total_time = time.time() - start_time
        
        excel.save()
        excel.close()
        await self.api_client.close()
        
        print(f"\n{'=' * 60}")
        print("导出完成")
        print(f"{'=' * 60}")
        print(f"  总数: {len(records)}")
        print(f"  成功: {progress['success']}")
        print(f"  失败: {progress['failed']}")
        print(f"  耗时: {total_time:.1f}s")
        print(f"  结果已写入: {input_path}")
        print(f"{'=' * 60}")


# ==================== 命令行入口 ====================

def main():
    parser = argparse.ArgumentParser(
        description="年运报告批量导出脚本（独立版，无需项目代码）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
依赖安装:
  pip install aiohttp pandas openpyxl

说明:
  脚本通过调用服务端 /annual-report/stream 接口生成年运报告。
  可选 --app-id 指定百炼智能体（如 2025 年新提示词），不传则用服务端配置。

示例:
  # 导出 2025 年运报告（并行3个用户）
  python3 annual_report_exporter.py \\
      --input ~/Desktop/内部测试.xlsx --year 2025 --concurrency 3 --verbose

  # 使用指定百炼智能体（2025 年新提示词）
  python3 annual_report_exporter.py \\
      --input ~/Desktop/内部测试.xlsx --year 2025 --app-id YOUR_BAILIAN_APP_ID --verbose

  # 仅处理第1行
  python3 annual_report_exporter.py \\
      --input ~/Desktop/内部测试.xlsx --year 2025 --row 1 --verbose
        """
    )
    
    parser.add_argument("--input", "-i", required=True, help="输入 Excel 文件路径")
    parser.add_argument("--year", "-y", type=int, default=2025, help="目标年份（默认: 2025）")
    parser.add_argument("--app-id", default=None,
                        help="百炼智能体 App ID（可选，用于指定新提示词智能体，如 2025 年运）")
    parser.add_argument("--concurrency", "-c", type=int, default=3,
                        help="最大并发用户数（默认: 3）")
    parser.add_argument("--base-url", default="http://8.210.52.217:8001",
                        help="API 基础 URL（默认: http://8.210.52.217:8001）")
    parser.add_argument("--row", "-r", type=int, default=None,
                        help="仅处理指定数据行（从1开始）")
    parser.add_argument("--verbose", "-v", action="store_true", help="详细输出")
    
    args = parser.parse_args()
    
    log_level = logging.DEBUG if args.verbose else logging.WARNING
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s [%(levelname)s] %(message)s"
    )
    
    exporter = AnnualReportExporter(
        base_url=args.base_url,
        year=args.year,
        app_id=args.app_id,
        verbose=args.verbose
    )
    
    asyncio.run(exporter.export(
        input_path=args.input,
        max_concurrency=args.concurrency,
        row=args.row
    ))


if __name__ == "__main__":
    main()
