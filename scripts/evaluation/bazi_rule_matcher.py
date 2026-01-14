#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
八字规则匹配评测脚本

基于八字和性别匹配指定类目下的规则，输出匹配和未匹配的规则ID及内容。
"""

import sys
import os
import asyncio
import argparse
from typing import Dict, Any, List, Optional
from datetime import datetime

# 添加项目根目录到路径
project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, project_root)

from scripts.evaluation.config import EvaluationConfig, DEFAULT_CONFIG
from scripts.evaluation.bazi_parser import BaziParser
from scripts.evaluation.bazi_to_solar import BaziToSolarConverter
from scripts.evaluation.rule_category_mapper import RuleCategoryMapper
from scripts.evaluation.rule_query_client import RuleQueryClient
from scripts.evaluation.api_client import BaziApiClient
from scripts.evaluation.data_parser import DataParser
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class BaziRuleMatcher:
    """八字规则匹配器"""
    
    def __init__(self, config: EvaluationConfig = None, verbose: bool = False):
        """
        初始化匹配器
        
        Args:
            config: 评测配置
            verbose: 是否显示详细进度
        """
        self.config = config or DEFAULT_CONFIG
        self.verbose = verbose
        self.api_client = BaziApiClient(self.config)
    
    def _log(self, message: str, level: str = "INFO"):
        """输出日志"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] [{level}] {message}")
    
    def _log_progress(self, message: str):
        """输出进度信息（仅verbose模式）"""
        if self.verbose:
            self._log(message, "PROGRESS")
    
    async def match_rules_for_row(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        为单行数据匹配规则
        
        Args:
            row_data: 行数据字典，包含：
                - bazi_text: 八字文本（换行分隔的4柱）
                - category_text: 匹配规则类目（空格分隔）
                - gender_text: 性别文本
                
        Returns:
            匹配结果字典，包含：
                - matched_rules: 匹配的规则列表
                - unmatched_rules: 未匹配的规则列表
                - error: 错误信息（如果有）
        """
        result = {
            'matched_rules': [],
            'unmatched_rules': [],
            'error': None
        }
        
        try:
            # 1. 解析八字
            bazi_text = row_data.get('bazi_text', '').strip()
            if not bazi_text or bazi_text == "nan":
                result['error'] = "八字文本为空"
                return result
            
            pillars = BaziParser.parse_bazi_lines(bazi_text)
            if not pillars:
                result['error'] = f"八字格式无效: {bazi_text}"
                return result
            
            # 2. 反推公历日期
            self._log_progress(f"  反推公历日期...")
            self._log_progress(f"  八字四柱: {pillars}")
            solar_result = BaziToSolarConverter.convert_bazi_to_solar(pillars)
            if not solar_result:
                result['error'] = f"无法反推公历日期，八字: {bazi_text}"
                if self.verbose:
                    # 详细错误信息
                    result['error'] += f"\n  年柱: {pillars.get('year', '')}, 月柱: {pillars.get('month', '')}, 日柱: {pillars.get('day', '')}, 时柱: {pillars.get('hour', '')}"
                return result
            
            solar_date, solar_time = solar_result
            self._log_progress(f"  公历日期: {solar_date} {solar_time}")
            
            # 3. 解析性别
            gender_text = row_data.get('gender_text', '').strip()
            try:
                gender = DataParser.parse_gender(gender_text)
            except ValueError as e:
                result['error'] = f"性别解析失败: {e}"
                return result
            
            # 4. 映射规则类目
            category_text = row_data.get('category_text', '').strip()
            rule_types = RuleCategoryMapper.map_categories_to_rule_types(category_text)
            if not rule_types:
                result['error'] = f"无法映射规则类目: {category_text}"
                return result
            
            self._log_progress(f"  规则类目: {category_text} -> {rule_types}")
            
            # 5. 调用API匹配规则
            self._log_progress(f"  调用规则匹配API...")
            try:
                api_result = await self.api_client.call_bazi_rules_match(
                    solar_date, solar_time, gender, rule_types
                )
                
                # 检查API返回的错误
                if api_result.get('error'):
                    result['error'] = f"API返回错误: {api_result['error']}"
                    return result
                
                # 检查success字段
                if not api_result.get('success', False):
                    result['error'] = f"API调用失败: {api_result.get('message', '未知错误')}"
                    return result
                
                matched_rules = api_result.get('matched_rules', [])
                self._log_progress(f"  匹配到 {len(matched_rules)} 条规则")
                
            except RuntimeError as e:
                # 捕获API调用异常（包括500错误）
                error_msg = str(e)
                if '500' in error_msg or 'Internal Server Error' in error_msg:
                    result['error'] = f"服务器内部错误(500)，可能原因：八字计算失败或规则匹配异常。请检查服务器日志。"
                    if self.verbose:
                        result['error'] += f"\n  详细错误: {error_msg}"
                else:
                    result['error'] = f"API调用失败: {error_msg}"
                return result
            except Exception as e:
                # 捕获其他异常
                result['error'] = f"API调用异常: {str(e)}"
                if self.verbose:
                    import traceback
                    result['error'] += f"\n  详细错误: {traceback.format_exc()}"
                return result
            
            # 6. 获取该类目下所有规则
            self._log_progress(f"  查询所有规则...")
            all_rules = RuleQueryClient.get_rules_by_type(rule_types)
            self._log_progress(f"  该类目下共有 {len(all_rules)} 条规则")
            
            # 7. 计算未匹配的规则
            matched_rule_codes = {rule.get('rule_code', '') for rule in matched_rules}
            unmatched_rules = [
                rule for rule in all_rules
                if rule.get('rule_code', '') not in matched_rule_codes
            ]
            self._log_progress(f"  未匹配 {len(unmatched_rules)} 条规则")
            
            result['matched_rules'] = matched_rules
            result['unmatched_rules'] = unmatched_rules
            
            return result
            
        except Exception as e:
            result['error'] = f"匹配失败: {e}"
            import traceback
            if self.verbose:
                traceback.print_exc()
            return result
    
    def format_rules_output(self, rules: List[Dict]) -> str:
        """
        格式化规则输出（ID + 内容）
        
        Args:
            rules: 规则列表
            
        Returns:
            格式化后的字符串
        """
        if not rules:
            return ""
        
        lines = []
        for rule in rules:
            rule_code = rule.get('rule_code', '')
            rule_name = rule.get('rule_name', '')
            
            # 提取规则内容文本
            content_text = RuleQueryClient.extract_rule_text(rule)
            
            lines.append(f"ID: {rule_code}")
            if rule_name:
                lines.append(f"名称: {rule_name}")
            if content_text:
                lines.append(f"内容: {content_text}")
            lines.append("")  # 空行分隔
        
        return "\n".join(lines)
    
    async def close(self):
        """关闭资源"""
        await self.api_client.close()


def read_excel_data(filepath: str, start_row: Optional[int] = None, 
                   end_row: Optional[int] = None) -> List[Dict[str, Any]]:
    """
    读取Excel数据
    
    Args:
        filepath: Excel文件路径
        start_row: 起始行号（从1开始，1表示第一条数据）
        end_row: 结束行号（包含）
        
    Returns:
        行数据列表
    """
    workbook = load_workbook(filepath)
    sheet = workbook.active
    
    rows = []
    
    # 假设第1行是表头，第2行开始是数据
    header_row = 1
    data_start_row = 2
    
    # 读取表头（确定列索引）
    headers = {}
    for col_idx, cell in enumerate(sheet[header_row], 1):
        headers[col_idx] = cell.value
    
    # 读取数据
    max_row = sheet.max_row
    for row_idx in range(data_start_row, max_row + 1):
        data_row_num = row_idx - data_start_row + 1  # 数据行号（从1开始）
        
        # 行号过滤
        if start_row is not None and data_row_num < start_row:
            continue
        if end_row is not None and data_row_num > end_row:
            break
        
        row_data = {}
        for col_idx in headers:
            cell = sheet.cell(row=row_idx, column=col_idx)
            row_data[col_idx] = cell.value
        
        # 根据实际Excel列位置映射（需要根据实际Excel格式调整）
        # 假设：A=标签, B=匹配规则类目, C=性别, D=八字, E=判词, F=匹配对应ID, G=不匹配对应ID
        bazi_text = str(row_data.get(4, '')).strip() if row_data.get(4) else ''
        
        # 处理八字文本：可能是换行分隔的4行，也可能是其他格式
        # 如果包含换行符，保持原样；否则尝试按其他格式解析
        if bazi_text and '\n' not in bazi_text:
            # 可能是空格或其他分隔符分隔的格式
            # 尝试按空格分割
            parts = bazi_text.split()
            if len(parts) == 4:
                bazi_text = '\n'.join(parts)
        
        mapped_data = {
            'label': str(row_data.get(1, '')).strip() if row_data.get(1) else '',  # A列：标签
            'category_text': str(row_data.get(2, '')).strip() if row_data.get(2) else '',  # B列：匹配规则类目
            'gender_text': str(row_data.get(3, '')).strip() if row_data.get(3) else '',  # C列：性别
            'bazi_text': bazi_text,  # D列：八字
            'judgment_text': str(row_data.get(5, '')).strip() if row_data.get(5) else '',  # E列：判词
            'row_index': row_idx,  # Excel行号
            'data_row_num': data_row_num  # 数据行号
        }
        
        rows.append(mapped_data)
    
    workbook.close()
    return rows


def write_excel_results(filepath: str, results: List[Dict[str, Any]]):
    """
    写入Excel结果
    
    Args:
        filepath: Excel文件路径
        results: 结果列表，每个元素包含：
            - row_index: Excel行号
            - matched_output: 匹配规则输出文本
            - unmatched_output: 未匹配规则输出文本
    """
    workbook = load_workbook(filepath)
    sheet = workbook.active
    
    for result in results:
        row_idx = result['row_index']
        matched_output = result.get('matched_output', '')
        unmatched_output = result.get('unmatched_output', '')
        
        # 写入F列：匹配对应ID
        sheet.cell(row=row_idx, column=6, value=matched_output)
        
        # 写入G列：不匹配对应ID
        sheet.cell(row=row_idx, column=7, value=unmatched_output)
    
    workbook.save(filepath)
    workbook.close()


async def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='八字规则匹配评测工具 - 基于八字和性别匹配规则',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 处理所有数据
  python bazi_rule_matcher.py --input /path/to/file.xlsx
  
  # 处理单条数据（第2行）
  python bazi_rule_matcher.py --input /path/to/file.xlsx --row 2
  
  # 显示详细进度
  python bazi_rule_matcher.py --input /path/to/file.xlsx --verbose
        """
    )
    
    parser.add_argument('--input', '-i', required=True, help='输入Excel文件路径')
    parser.add_argument('--row', '-r', type=int, help='只处理指定数据行（从1开始，1表示第一条数据）')
    parser.add_argument('--verbose', '-v', action='store_true', help='显示详细进度')
    parser.add_argument('--base-url', help='API服务地址（默认: http://8.210.52.217:8001）')
    
    args = parser.parse_args()
    
    # 创建配置
    config = EvaluationConfig()
    if args.base_url:
        config.base_url = args.base_url
    
    # 创建匹配器
    matcher = BaziRuleMatcher(config=config, verbose=args.verbose)
    
    try:
        # 读取Excel数据
        print(f"读取Excel文件: {args.input}")
        rows = read_excel_data(
            args.input,
            start_row=args.row if args.row else None,
            end_row=args.row if args.row else None
        )
        
        if not rows:
            print("未找到有效的测试数据")
            return
        
        print(f"读取到 {len(rows)} 条数据")
        
        # 处理每一行
        results = []
        for i, row_data in enumerate(rows, 1):
            print(f"\n处理第 {i}/{len(rows)} 行（数据行号: {row_data['data_row_num']}）...")
            
            match_result = await matcher.match_rules_for_row(row_data)
            
            if match_result.get('error'):
                print(f"错误: {match_result['error']}")
                continue
            
            # 格式化输出
            matched_output = matcher.format_rules_output(match_result['matched_rules'])
            unmatched_output = matcher.format_rules_output(match_result['unmatched_rules'])
            
            results.append({
                'row_index': row_data['row_index'],
                'matched_output': matched_output,
                'unmatched_output': unmatched_output
            })
        
        # 写入结果
        print("\n写入结果到Excel...")
        write_excel_results(args.input, results)
        print(f"完成！结果已保存到: {args.input}")
        
    except FileNotFoundError as e:
        print(f"错误: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"处理过程中出错: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        await matcher.close()


if __name__ == "__main__":
    asyncio.run(main())
