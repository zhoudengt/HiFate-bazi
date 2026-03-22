#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
评测 Excel 读写处理

读取测试用例、写入评测结果，供 bazi_evaluator 使用。
"""

import os
from dataclasses import dataclass
from typing import List, Optional, Any, Dict

import pandas as pd
from openpyxl import load_workbook

from scripts.evaluation.config import ExcelColumns
from scripts.evaluation.data_parser import DataParser


@dataclass
class TestCase:
    """测试用例"""
    user_name: str
    solar_date: str
    solar_time: str
    gender: str
    day_pillar: str = ""
    bazi_text: str = ""
    parse_error: Optional[str] = None
    row_index: int = 0


class ExcelHandler:
    """评测 Excel 读写处理器"""

    def __init__(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Excel 文件不存在: {filepath}")
        self.filepath = filepath
        self._workbook = load_workbook(filepath)
        self._sheet = self._workbook.active
        self._df = pd.read_excel(filepath, header=0)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False

    def read_test_cases(
        self,
        data_row_start: Optional[int] = None,
        data_row_end: Optional[int] = None
    ) -> List[TestCase]:
        """
        读取测试用例（Excel 第 1 行为表头，数据从第 2 行起）

        data_row_start/end: 从 1 开始，1 表示第一条数据行
        """
        cases = []
        for idx, row in self._df.iterrows():
            data_row_num = idx + 1  # 数据行号从1开始，1=第一条数据
            if data_row_start is not None and data_row_num < data_row_start:
                continue
            if data_row_end is not None and data_row_num > data_row_end:
                break

            user_name = _cell(row, ExcelColumns.USER_NAME)
            birth_text = _cell(row, ExcelColumns.USER_BIRTH)
            gender_text = _cell(row, ExcelColumns.GENDER)
            bazi_text = _cell(row, ExcelColumns.BAZI)
            day_pillar = ""

            if not birth_text or str(birth_text).strip() == "" or str(birth_text) == "nan":
                continue

            # 跳过表头行（birth 或 gender 列包含表头常见字样）
            birth_str = str(birth_text).strip().lower()
            gender_str = str(gender_text or "").strip().lower()
            _header_keywords = ('birth', '生辰', 'user_', '日期', '性别', 'gender', '男/女')
            if any(kw in birth_str for kw in _header_keywords) or any(kw in gender_str for kw in _header_keywords):
                continue

            solar_date, solar_time, gender = "", "", ""
            parse_error = None

            try:
                solar_date, solar_time = DataParser.parse_birth_with_time(birth_text)
            except ValueError as e:
                parse_error = str(e)

            try:
                gender = DataParser.parse_gender(gender_text) if gender_text else ""
            except ValueError as e:
                parse_error = f"{parse_error}; {e}" if parse_error else str(e)

            # 从八字列提取日柱（如 "甲子" 格式的日柱）
            if bazi_text and str(bazi_text) != "nan":
                parts = str(bazi_text).replace('\n', ' ').split()
                if len(parts) >= 3:  # 年 月 日 时，日柱为第3个
                    day_pillar = parts[2].strip()

            cases.append(TestCase(
                user_name=user_name or "",
                solar_date=solar_date,
                solar_time=solar_time,
                gender=gender,
                day_pillar=day_pillar,
                bazi_text=bazi_text or "",
                parse_error=parse_error,
                row_index=idx
            ))
        return cases

    def write_result(self, test_case: TestCase, result: Dict[str, Any]) -> None:
        """写入评测结果到对应行"""
        excel_row = test_case.row_index + 2  # 1=表头

        # 基础数据
        basic = result.get('basic', {})
        if basic:
            self._write_cell(excel_row, ExcelColumns.DAY_STEM, basic.get('day_stem', ''))
            self._write_cell(excel_row, ExcelColumns.WUXING, basic.get('wuxing', ''))
            self._write_cell(excel_row, ExcelColumns.SHISHEN, basic.get('shishen', ''))
            self._write_cell(excel_row, ExcelColumns.XI_JI, basic.get('xi_ji', ''))
            self._write_cell(excel_row, ExcelColumns.WANGSHUAI, basic.get('wangshuai', ''))
            self._write_cell(excel_row, ExcelColumns.GEJU, basic.get('geju', ''))
            self._write_cell(excel_row, ExcelColumns.DAYUN_LIUNIAN, basic.get('dayun_liunian', ''))

        # 日元六十甲子 + LLM 分析（从流式接口 content 提取）
        self._write_cell(excel_row, ExcelColumns.RIZHU_LIUJIAZI, result.get('rizhu_liujiazi', ''))
        self._write_cell(excel_row, ExcelColumns.WUXING_ANALYSIS, result.get('wuxing_analysis', ''))
        self._write_cell(excel_row, ExcelColumns.XISHEN_JISHEN, result.get('xishen_jishen', ''))
        self._write_cell(excel_row, ExcelColumns.CAREER_WEALTH, result.get('career_wealth', ''))
        self._write_cell(excel_row, ExcelColumns.MARRIAGE, result.get('marriage', ''))
        self._write_cell(excel_row, ExcelColumns.HEALTH, result.get('health', ''))
        self._write_cell(excel_row, ExcelColumns.CHILDREN_STUDY, result.get('children_study', ''))
        self._write_cell(excel_row, ExcelColumns.GENERAL_REVIEW, result.get('general_review', ''))
        self._write_cell(excel_row, ExcelColumns.DAILY_FORTUNE, result.get('daily_fortune', ''))
        self._write_cell(excel_row, ExcelColumns.YEAR_FORTUNE, result.get('annual_report', ''))

    def _write_cell(self, row: int, col: int, value: Any) -> None:
        """写入单元格（openpyxl 行列从 1 开始）"""
        c = self._sheet.cell(row=row, column=col + 1, value=_str_value(value))

    def save(self) -> None:
        self._workbook.save(self.filepath)

    def close(self) -> None:
        if self._workbook:
            self._workbook.close()
            self._workbook = None


def _cell(row, col: int) -> str:
    v = row.iloc[col] if col < len(row) else None
    return "" if (v is None or (hasattr(v, '__iter__') and not isinstance(v, str)) and pd.isna(v)) else str(v)


def _str_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    return str(v)
